import io, os, pdfplumber
from datetime import datetime
from flask import Flask, request, send_file, render_template_string
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Import parser lokal
from bca_parser import parse_bca
from bri_parser import parse_bri
from mandiri_parser import parse_mandiri

app = Flask(__name__)

def classify_common(text, is_credit):
    u = text.upper()
    m = {'PENERIMAAN NEGARA':'', 'PAJAK':'29', 'BUNGA':'28', 'BIAYA ADM':'27', 'TRANSFER':'27', 'TELKOM':'27', 'LISTRIK':'27', 'GAJI':''}
    for k, v in m.items():
        if k in u: return k, v
    if 'SETORAN TUNAI' in u: return 'SETORAN TUNAI', '1'
    return ('PENERIMAAN PENJUALAN', '3') if is_credit else ('PELUNASAN HUTANG DAGANG', '')

def detect_bank(all_lines):
    text = ' '.join(all_lines[:30]).upper()
    if any(x in text for x in ['REKENING GIRO', 'TRSF E-BANKING', 'BI-FAST']): return 'BCA'
    if any(x in text for x in ['BRIMTXDT', 'BRITAMA', 'LAPORAN TRANSAKSI FINANSIAL']): return 'BRI'
    if any(x in text for x in ['KOPRA', 'MANDIRI', 'MCM INHOUSETRF']): return 'MANDIRI'
    return 'UNKNOWN'

def create_excel(data):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Mutasi {data['bank']}"
    colors = {'BCA': '005BAC', 'BRI': '003D7C', 'MANDIRI': '003087'}.get(data['bank'], '333333')
    
    ws.merge_cells('A1:I1'); ws['A1'] = f"LAPORAN MUTASI – {data['bank']}"
    ws['A1'].font = Font(bold=True, color='FFFFFF', size=12); ws['A1'].fill = PatternFill('solid', start_color=colors)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    headers = ['NO', 'TANGGAL', 'NAMA', 'KETERANGAN', 'KODE', 'DEBET', 'KREDIT', 'SALDO', 'PENJELASAN']
    ws.append(headers)
    for c in range(1, 10):
        ws.cell(2, c).font = Font(bold=True, color='FFFFFF')
        ws.cell(2, c).fill = PatternFill('solid', start_color=colors)

    saldo = data['saldo_awal']
    for i, tx in enumerate(data['txs'], 1):
        saldo = round(saldo + tx['kredit'] - tx['debet'], 2)
        ws.append([i, tx['date'], tx['nama'], tx['ket'], tx['kode'], tx['debet'] or '', tx['kredit'] or '', saldo, tx['penjelasan']])

    out = io.BytesIO()
    wb.save(out); out.seek(0)
    return out

@app.route('/')
def index():
    # Gunakan template HTML dari Source 5 (Dark Mode Modern)
    return render_template_string(open(os.path.join(os.path.dirname(__file__), 'template.html')).read() if os.path.exists('template.html') else "File template.html tidak ditemukan. Gunakan UI dari Source 5.")

@app.route('/convert', methods=['POST'])
def convert():
    files = request.files.getlist('files')
    all_lines = []
    with pdfplumber.open(io.BytesIO(files[0].read())) as pdf:
        for p in pdf.pages: all_lines.extend(p.extract_text().split('\n'))
    
    bank = detect_bank(all_lines)
    if bank == 'BCA': data = parse_bca(all_lines)
    elif bank == 'BRI': data = parse_bri(all_lines, classify_common)
    elif bank == 'MANDIRI': data = parse_mandiri(all_lines, classify_common)
    else: return "Bank tidak dikenali", 400

    excel = create_excel(data)
    return send_file(excel, as_attachment=True, download_name=f"Mutasi_{bank}.xlsx")

if __name__ == '__main__':
    app.run(debug=True, port=5000)