"""
Bank Statement PDF → Excel Converter
Mendukung: BCA, BRI, Mandiri
Versi: 2.0
"""

import re
import io
from datetime import datetime

import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from flask import Flask, request, send_file, render_template_string

app = Flask(__name__)

# ─────────────────────────────────────────────
# KLASIFIKASI KETERANGAN & KODE
# ─────────────────────────────────────────────

KODE_MAP = {
    'PENERIMAAN PENJUALAN':     '5',
    'SETORAN TUNAI':            '1',
    'PELUNASAN PIUTANG DAGANG': '1',
    'PELUNASAN HUTANG DAGANG':  '',
    'BIAYA ADM BANK':           '27',
    'BIAYA TRANSFER':           '27',
    'BIAYA TELEPON':            '27',
    'BIAYA LISTRIK':            '27',
    'BIAYA GAJI PEGAWAI':       '',
    'BUNGA':                    '28',
    'PAJAK JASA GIRO':          '29',
    'PENERIMAAN NEGARA':        '',
    'KOREKSI BANK':             '',
}

def classify(text: str, is_credit: bool) -> tuple[str, str]:
    u = text.upper()
    if 'PENERIMAAN NEGARA' in u:                        ket = 'PENERIMAAN NEGARA'
    elif 'PAJAK' in u:                                  ket = 'PAJAK JASA GIRO'
    elif 'BUNGA' in u or 'INTEREST' in u:               ket = 'BUNGA'
    elif 'BIAYA ADM' in u or 'BIAYA ADM' in u:          ket = 'BIAYA ADM BANK'
    elif 'TRANSFER' in u and 'BIAYA' in u:              ket = 'BIAYA TRANSFER'
    elif 'TELKOM' in u or 'TELEPON' in u:               ket = 'BIAYA TELEPON'
    elif 'LISTRIK' in u or 'PLN' in u:                  ket = 'BIAYA LISTRIK'
    elif 'GAJI' in u or 'SALARY' in u:                  ket = 'BIAYA GAJI PEGAWAI'
    elif 'SETORAN TUNAI' in u:                          ket = 'SETORAN TUNAI'
    elif 'KOREKSI' in u:                                ket = 'KOREKSI BANK'
    elif not is_credit:                                 ket = 'PELUNASAN HUTANG DAGANG'
    else:                                               ket = 'PENERIMAAN PENJUALAN'
    return ket, KODE_MAP.get(ket, '')


# ─────────────────────────────────────────────
# DETEKTOR JENIS BANK
# ─────────────────────────────────────────────

def detect_bank(all_lines: list[str]) -> str:
    text = ' '.join(all_lines[:30]).upper()
    if 'REKENING GIRO' in text or 'TRSF E-BANKING' in text or 'BI-FAST' in text:
        return 'BCA'
    if 'LAPORAN TRANSAKSI FINANSIAL' in text or 'BRIMTXDT' in text or 'BRITAMA' in text:
        return 'BRI'
    if 'KOPRA' in text or 'MANDIRI' in text or 'MCM INHOUSETRF' in text or 'ACCOUNT STATEMENT' in text:
        return 'MANDIRI'
    return 'UNKNOWN'


# ─────────────────────────────────────────────
# PARSER BCA
# Format: DD/MM TYPE CR/DB DESC  AMT  [SALDO]
# ─────────────────────────────────────────────

def parse_bca(all_lines: list[str]) -> dict:
    # Header info
    period = ''
    saldo_awal = saldo_akhir = mut_cr = mut_db = 0
    no_rek = ''

    for l in all_lines:
        if 'PERIODE :' in l.upper() and not period:
            period = l.split(':', 1)[1].strip()
        if 'NO. REKENING' in l.upper() and not no_rek:
            m = re.search(r':\s*([\d]+)', l)
            if m: no_rek = m.group(1)
        m = re.search(r'SALDO AWAL\s*:\s*([\d,]+\.\d+)', l.upper())
        if m: saldo_awal = float(m.group(1).replace(',', ''))
        m = re.search(r'MUTASI CR\s*:\s*([\d,]+\.\d+)', l.upper())
        if m: mut_cr = float(m.group(1).replace(',', ''))
        m = re.search(r'MUTASI DB\s*:\s*([\d,]+\.\d+)', l.upper())
        if m: mut_db = float(m.group(1).replace(',', ''))
        m = re.search(r'SALDO AKHIR\s*:\s*([\d,]+\.\d+)', l.upper())
        if m: saldo_akhir = float(m.group(1).replace(',', ''))

    year = period.split()[-1] if period else str(datetime.now().year)

    # Transaction parsing
    DATE_RE = re.compile(r'^(\d{2})/(\d{2})\s+(.+)')
    SKIP_STARTS = ('REKENING GIRO', 'KCU ', 'MITRA', 'BOJONGLOA', 'SITUSAEUR',
                   'JL ', 'BANDUNG', 'INDONESIA', 'NO. REKENING', 'HALAMAN',
                   'PERIODE', 'MATA UANG', 'CATATAN', 'APABILA', 'REKENING INI',
                   'TELAH MENYETUJUI', 'BCA BERHAK', 'LAPORAN MUTASI',
                   'TANGGAL KETERANGAN', 'SALDO AWAL :', 'MUTASI CR', 'MUTASI DB',
                   'SALDO AKHIR', 'BERSAMBUNG')

    txs = []
    for l in all_lines:
        l = l.strip()
        dm = DATE_RE.match(l)
        if not dm:
            continue
        day, mon, rest = dm.group(1), dm.group(2), dm.group(3).strip()
        if rest == 'SALDO AWAL':
            continue
        if any(rest.upper().startswith(s) for s in SKIP_STARTS):
            continue

        all_amts = re.findall(r'([\d,]+\.\d{2})', l)
        if not all_amts:
            continue

        amt = float(all_amts[0].replace(',', ''))
        if amt <= 0:
            continue

        u = rest.upper()
        is_cr = (' CR ' in u or 'SETORAN TUNAI' in u or
                 'KR OTOMATIS' in u or 'SWITCHING CR' in u)
        is_db = (' DB ' in u or 'BYR VIA' in u)

        # Override for special transaction types
        if 'BIAYA ADM' in u or 'PAJAK' in u:
            is_db, is_cr = True, False
        elif u.strip().startswith('BUNGA'):
            is_db, is_cr = False, True

        ket, kode = classify(rest, is_cr and not is_db)
        date_str = f"{day}/{mon}/{year[-2:]}"

        txs.append({
            'date': date_str,
            'ket':  ket,
            'kode': kode,
            'debet':  amt if is_db else 0.0,
            'kredit': amt if (is_cr and not is_db) else 0.0,
        })

    # Fallback: if kredit+debet both 0, assign as kredit
    for tx in txs:
        if tx['debet'] == 0 and tx['kredit'] == 0:
            tx['kredit'] = next(
                (float(a.replace(',', ''))
                 for a in re.findall(r'([\d,]+\.\d{2})', tx.get('_raw', ''))
                 if float(a.replace(',', '')) > 0),
                0.0
            )

    nama_akun = 'CV. MITRA JAYA ANUGERAH'
    for l in all_lines[:15]:
        if 'CV.' in l.upper() or 'PT.' in l.upper():
            nama_akun = l.strip()
            break

    return {
        'bank':       'BCA',
        'no_rek':     no_rek,
        'nama_akun':  nama_akun,
        'period':     period,
        'saldo_awal': saldo_awal,
        'saldo_akhir':saldo_akhir,
        'mut_cr':     mut_cr,
        'mut_db':     mut_db,
        'txs':        txs,
    }


# ─────────────────────────────────────────────
# PARSER BRI
# Format: DD/MM/YY HH:MM:SS DESC TELLER DB CR SALDO
# ─────────────────────────────────────────────

def parse_bri(all_lines: list[str]) -> dict:
    period = ''
    saldo_awal = saldo_akhir = total_db = total_cr = 0
    no_rek = ''
    nama_akun = ''

    for l in all_lines:
        # Period
        m = re.search(r'(?:Periode Transaksi|Transaction Period)[^\d]*(\d{2}/\d{2}/\d{2})\s*-\s*(\d{2}/\d{2}/\d{2})', l)
        if m and not period:
            period = f"{m.group(1)} - {m.group(2)}"
        # Account number
        m = re.search(r'No\.\s*Rekening[^\d]*([\d]+)', l)
        if m and not no_rek:
            no_rek = m.group(1)
        # Company name
        if l.strip().startswith('CV ') or l.strip().startswith('PT '):
            if not nama_akun:
                nama_akun = l.strip()
        # Summary: saldo_awal total_db total_cr saldo_akhir (4 numbers on one line)
        m = re.match(r'^([\d,]+\.\d{2})\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})$', l.strip())
        if m:
            saldo_awal  = float(m.group(1).replace(',', ''))
            total_db    = float(m.group(2).replace(',', ''))
            total_cr    = float(m.group(3).replace(',', ''))
            saldo_akhir = float(m.group(4).replace(',', ''))

    # Transaction line: DD/MM/YY HH:MM:SS DESC TELLER? 0.00 0.00 0.00
    TX_RE = re.compile(
        r'^(\d{2}/\d{2}/\d{2})\s+\d{2}:\d{2}:\d{2}\s+'
        r'(.+?)\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})$'
    )

    txs = []
    for l in all_lines:
        m = TX_RE.match(l.strip())
        if not m:
            continue
        date_raw, desc, db_s, cr_s, _ = m.groups()
        db  = float(db_s.replace(',', ''))
        cr  = float(cr_s.replace(',', ''))
        is_credit = cr > 0
        ket, kode = classify(desc, is_credit)
        txs.append({
            'date':   date_raw,
            'ket':    ket,
            'kode':   kode,
            'debet':  db,
            'kredit': cr,
        })

    return {
        'bank':       'BRI',
        'no_rek':     no_rek,
        'nama_akun':  nama_akun or 'CV. TUNGGAL JAYA SUNIARAJA',
        'period':     period,
        'saldo_awal': saldo_awal,
        'saldo_akhir':saldo_akhir,
        'mut_cr':     total_cr,
        'mut_db':     total_db,
        'txs':        txs,
    }


# ─────────────────────────────────────────────
# PARSER MANDIRI (Kopra)
# Format: multi-line; amounts always end each transaction line
# Pattern: ...TEXT... DEBIT CREDIT BALANCE
# ─────────────────────────────────────────────

_MANDIRI_MONTH = {
    'Jan':'01','Feb':'02','Mar':'03','Apr':'04','May':'05','Jun':'06',
    'Jul':'07','Aug':'08','Sep':'09','Oct':'10','Nov':'11','Dec':'12',
}

def parse_mandiri(all_lines: list[str]) -> dict:
    period = ''
    saldo_awal = saldo_akhir = total_db = total_cr = 0
    no_rek = ''
    nama_akun = ''

    for l in all_lines:
        m = re.search(r'(\d{2}\s+\w+\s+\d{4})\s*-\s*(\d{2}\s+\w+\s+\d{4})', l)
        if m and not period:
            period = f"{m.group(1)} - {m.group(2)}"
        m = re.match(r'^(\d{10,16})\s+(.+)', l.strip())
        if m and not no_rek:
            no_rek    = m.group(1)
            nama_akun = m.group(2).split('  ')[0].strip()
        # "99,678,458.14 4 158,231,858.44" → saldo_awal + count + total_db
        # "44,529,166.88 16 103,082,567.18" → saldo_akhir + count + total_cr
        m = re.match(r'^([\d,]+\.\d{2})\s+\d+\s+([\d,]+\.\d{2})$', l.strip())
        if m:
            v1 = float(m.group(1).replace(',', ''))
            v2 = float(m.group(2).replace(',', ''))
            if saldo_awal == 0:
                saldo_awal = v1; total_db = v2
            else:
                saldo_akhir = v1; total_cr = v2

    # Transaction line detector: ends with 3 numbers (debit credit balance)
    TX_AMT_RE = re.compile(
        r'^(.*?)\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})\s*$'
    )
    DATE_RE = re.compile(
        r'(\d{2})\s+(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+(\d{4})',
        re.IGNORECASE
    )

    SKIP = ('For further questions', 'Account Statement', 'Created', 'Posting Date',
            'Opening Balance', 'Closing Balance', 'No. of Debit', 'Total Amount',
            'Account Statement Summary', 'Account No.', 'Period ', 'Alias',
            'Currency', 'Branch', 'kopra')

    txs = []
    last_date = ''

    for i, l in enumerate(all_lines):
        ls = l.strip()
        if not ls or any(ls.lower().startswith(s.lower()) for s in SKIP):
            continue

        # Track date
        dm = DATE_RE.search(ls)
        if dm:
            day = dm.group(1)
            mon = _MANDIRI_MONTH.get(dm.group(2)[:3].capitalize(), '01')
            yr  = dm.group(3)[-2:]
            last_date = f"{day}/{mon}/{yr}"

        # Transaction line: ends with 3 numbers
        m = TX_AMT_RE.match(ls)
        if not m:
            continue

        prefix = m.group(1).strip()
        db     = float(m.group(2).replace(',', ''))
        cr     = float(m.group(3).replace(',', ''))
        # m.group(4) is balance — not stored, recalculated later

        # Skip balance-only summary rows
        if any(prefix.lower().startswith(s.lower()) for s in
               ('Closing', 'Opening', 'Total', 'Terbilang')):
            continue

        # Build description: clean prefix + nearby lines
        desc = re.sub(r'\d{2}:\d{2}:\d{2}', '', prefix)   # remove time
        desc = re.sub(r'\b\d{8,}\b', '', desc)             # remove long ref nums
        desc = re.sub(r'\s{2,}', ' ', desc).strip(' -')

        # Pull in neighboring description lines
        for offset in [-2, -1, 1, 2]:
            ni = i + offset
            if not (0 <= ni < len(all_lines)):
                continue
            nb = all_lines[ni].strip()
            if (nb and not TX_AMT_RE.match(nb)
                    and not DATE_RE.search(nb)
                    and not any(nb.lower().startswith(s.lower()) for s in SKIP)
                    and not re.match(r'^\d{2}:\d{2}:\d{2}', nb)):
                desc = (desc + ' ' + nb).strip()
                break

        if not last_date:
            continue

        is_credit = cr > 0
        ket, kode = classify(desc or ls, is_credit)

        txs.append({'date': last_date, 'ket': ket, 'kode': kode,
                    'debet': db, 'kredit': cr})

    return {
        'bank':        'MANDIRI',
        'no_rek':      no_rek,
        'nama_akun':   nama_akun or 'TUNGGAL JAYA SUNIARA',
        'period':      period,
        'saldo_awal':  saldo_awal,
        'saldo_akhir': saldo_akhir,
        'mut_cr':      total_cr,
        'mut_db':      total_db,
        'txs':         txs,
    }


# ─────────────────────────────────────────────
# DISPATCHER UTAMA
# ─────────────────────────────────────────────

def parse_pdf(pdf_bytes: bytes) -> dict:
    all_lines = []
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if text:
                all_lines.extend(text.split('\n'))

    bank = detect_bank(all_lines)

    if bank == 'BCA':
        return parse_bca(all_lines)
    elif bank == 'BRI':
        return parse_bri(all_lines)
    elif bank == 'MANDIRI':
        return parse_mandiri(all_lines)
    else:
        raise ValueError(f"Format bank tidak dikenali. Pastikan PDF adalah mutasi BCA, BRI, atau Mandiri.")


# ─────────────────────────────────────────────
# GENERATOR EXCEL
# ─────────────────────────────────────────────

BANK_COLORS = {
    'BCA':     {'header': '005BAC', 'sub': 'E8F0FA', 'accent': '0072CE'},
    'BRI':     {'header': '003D7C', 'sub': 'E6EEF7', 'accent': '0055B3'},
    'MANDIRI': {'header': '003087', 'sub': 'E5ECF6', 'accent': 'F5A623'},
}

def create_excel(data: dict) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Mutasi {data['bank']}"

    bank   = data['bank']
    colors = BANK_COLORS.get(bank, BANK_COLORS['BCA'])
    NUM    = '#,##0.00'

    def style_cell(cell, bold=False, color=None, bg=None, align='left', num_fmt=None):
        cell.font      = Font(name='Calibri', bold=bold, color=color or '000000', size=10)
        cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=False)
        if bg:
            cell.fill = PatternFill('solid', start_color=bg)
        if num_fmt:
            cell.number_format = num_fmt

    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Title block ──
    ws.merge_cells('A1:I1')
    ws['A1'] = f"LAPORAN MUTASI REKENING – {bank}"
    style_cell(ws['A1'], bold=True, color='FFFFFF', bg=colors['header'], align='center')
    ws['A1'].font = Font(name='Calibri', bold=True, color='FFFFFF', size=13)
    ws.row_dimensions[1].height = 26

    ws.merge_cells('A2:I2')
    ws['A2'] = data['nama_akun']
    style_cell(ws['A2'], bold=True, align='center', bg=colors['sub'])
    ws.row_dimensions[2].height = 18

    ws.merge_cells('A3:I3')
    ws['A3'] = f"No. Rekening: {data['no_rek']}    |    Periode: {data['period']}"
    style_cell(ws['A3'], align='center', bg=colors['sub'])
    ws.row_dimensions[3].height = 16

    # ── Summary row ──
    ws.row_dimensions[4].height = 14
    for c in ['A','B','C','D','E','F','G','H','I']:
        ws[f'{c}4'].fill = PatternFill('solid', start_color='F5F5F5')

    summaries = [
        ('Saldo Awal', data['saldo_awal']),
        ('Total Kredit', data['mut_cr']),
        ('Total Debet', data['mut_db']),
        ('Saldo Akhir', data['saldo_akhir']),
    ]
    cols = [1, 3, 5, 7]
    for (label, val), col in zip(summaries, cols):
        lc = ws.cell(5, col, label)
        style_cell(lc, bold=True, align='right', bg='F0F4FF')
        ws.merge_cells(start_row=5, start_column=col, end_row=5, end_column=col+1)
        vc = ws.cell(5, col+2, val)
        style_cell(vc, bold=True, align='right', bg='FAFCFF', num_fmt=NUM)
    ws.row_dimensions[5].height = 18

    # ── Header row ──
    headers = ['NO', 'TANGGAL', 'KETERANGAN', 'KODE', 'DEBET', 'KREDIT', 'SALDO', '', '']
    header_cols = ['A','B','C','D','E','F','G']
    ws.row_dimensions[6].height = 20
    for col_letter, hdr in zip(header_cols, headers[:7]):
        c = ws[f'{col_letter}6']
        c.value = hdr
        style_cell(c, bold=True, color='FFFFFF', bg=colors['header'], align='center')
        c.border = border

    # ── Data rows ──
    saldo = data['saldo_awal']
    for idx, tx in enumerate(data['txs'], 1):
        saldo = round(saldo + tx['kredit'] - tx['debet'], 2)
        row = idx + 6
        bg = 'FFFFFF' if idx % 2 == 0 else 'F8FAFF'

        values = [idx, tx['date'], tx['ket'], tx['kode'],
                  tx['debet'] or '', tx['kredit'] or '', saldo]

        for col_idx, val in enumerate(values, 1):
            c = ws.cell(row, col_idx, val)
            num = None
            aln = 'center'
            if col_idx == 1:   aln = 'center'
            elif col_idx == 2: aln = 'center'
            elif col_idx == 3: aln = 'left'
            elif col_idx == 4: aln = 'center'
            elif col_idx in (5, 6, 7):
                aln = 'right'
                if isinstance(val, float) and val > 0:
                    num = NUM
            style_cell(c, align=aln, bg=bg, num_fmt=num)
            c.border = border
        ws.row_dimensions[row].height = 15

    # ── Totals row ──
    total_row = len(data['txs']) + 7
    ws.row_dimensions[total_row].height = 18
    labels = ['', 'TOTAL', '', '', data['mut_db'], data['mut_cr'], data['saldo_akhir']]
    for ci, val in enumerate(labels, 1):
        c = ws.cell(total_row, ci, val)
        num = NUM if ci in (5, 6, 7) and isinstance(val, float) else None
        aln = 'right' if ci in (5, 6, 7) else ('center' if ci == 2 else 'left')
        style_cell(c, bold=True, bg=colors['sub'], align=aln, num_fmt=num)
        c.border = border

    # ── Column widths ──
    col_widths = {'A': 6, 'B': 12, 'C': 30, 'D': 7, 'E': 16, 'F': 16, 'G': 18}
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    # ── Freeze panes ──
    ws.freeze_panes = 'A7'

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


# ─────────────────────────────────────────────
# FLASK ROUTES
# ─────────────────────────────────────────────

HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="id">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0, maximum-scale=1.0">
<title>Bank Statement Converter</title>
<style>
  @import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Sans:wght@400;500;600;700&family=IBM+Plex+Mono:wght@400;500&display=swap');

  :root {
    --bg:       #0a0e1a;
    --surface:  #111827;
    --card:     #161d2e;
    --border:   #1e2d45;
    --accent:   #3b82f6;
    --accent2:  #06b6d4;
    --text:     #e2e8f0;
    --muted:    #64748b;
    --success:  #10b981;
    --danger:   #ef4444;
    --radius:   16px;
  }

  *, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }

  body {
    font-family: 'IBM Plex Sans', sans-serif;
    background: var(--bg);
    color: var(--text);
    min-height: 100vh;
    display: flex;
    flex-direction: column;
    align-items: center;
    padding: 24px 16px 48px;
  }

  /* Grid background */
  body::before {
    content: '';
    position: fixed;
    inset: 0;
    background-image:
      linear-gradient(rgba(59,130,246,.04) 1px, transparent 1px),
      linear-gradient(90deg, rgba(59,130,246,.04) 1px, transparent 1px);
    background-size: 40px 40px;
    pointer-events: none;
    z-index: 0;
  }

  .container { position: relative; z-index: 1; width: 100%; max-width: 520px; }

  /* Header */
  .header { text-align: center; margin-bottom: 32px; }
  .logo-wrap {
    display: inline-flex; align-items: center; justify-content: center;
    width: 64px; height: 64px;
    background: linear-gradient(135deg, var(--accent), var(--accent2));
    border-radius: 18px; margin-bottom: 16px;
    box-shadow: 0 0 32px rgba(59,130,246,.35);
  }
  .logo-wrap svg { width: 32px; height: 32px; color: white; }
  h1 { font-size: clamp(20px,5vw,28px); font-weight: 700; letter-spacing: -0.5px; }
  .subtitle { color: var(--muted); font-size: 14px; margin-top: 6px; }

  /* Bank badges */
  .banks {
    display: flex; gap: 8px; justify-content: center; margin-top: 16px; flex-wrap: wrap;
  }
  .badge {
    font-family: 'IBM Plex Mono', monospace;
    font-size: 11px; font-weight: 500;
    padding: 4px 10px; border-radius: 6px; border: 1px solid;
  }
  .badge-bca    { color: #60a5fa; border-color: #1e3a5f; background: #0a1929; }
  .badge-bri    { color: #34d399; border-color: #1a3a2e; background: #0a1f18; }
  .badge-mandiri{ color: #fbbf24; border-color: #3a2f0a; background: #1f1800; }

  /* Card */
  .card {
    background: var(--card);
    border: 1px solid var(--border);
    border-radius: var(--radius);
    overflow: hidden;
    margin-bottom: 16px;
  }
  .card-header {
    background: var(--surface);
    border-bottom: 1px solid var(--border);
    padding: 14px 20px;
    font-size: 12px; font-weight: 600;
    text-transform: uppercase; letter-spacing: .08em;
    color: var(--muted);
    display: flex; align-items: center; gap: 8px;
  }
  .card-header::before {
    content: '';
    display: block; width: 3px; height: 14px;
    background: linear-gradient(var(--accent), var(--accent2));
    border-radius: 2px;
  }
  .card-body { padding: 20px; }

  /* Drop zone */
  .drop-zone {
    border: 2px dashed var(--border);
    border-radius: 12px;
    padding: 32px 20px;
    text-align: center;
    cursor: pointer;
    transition: all .2s;
    background: rgba(59,130,246,.03);
    position: relative;
  }
  .drop-zone:hover, .drop-zone.dragover {
    border-color: var(--accent);
    background: rgba(59,130,246,.08);
  }
  .drop-zone input[type=file] {
    position: absolute; inset: 0; opacity: 0; cursor: pointer;
  }
  .drop-icon {
    width: 44px; height: 44px;
    background: linear-gradient(135deg, rgba(59,130,246,.2), rgba(6,182,212,.2));
    border-radius: 12px;
    display: flex; align-items: center; justify-content: center;
    margin: 0 auto 12px;
  }
  .drop-icon svg { width: 22px; height: 22px; color: var(--accent); }
  .drop-label { font-size: 14px; color: var(--muted); line-height: 1.5; }
  .drop-label strong { color: var(--text); }
  .file-list { margin-top: 16px; display: flex; flex-direction: column; gap: 8px; }
  .file-item {
    display: flex; align-items: center; gap: 10px;
    background: var(--surface); border: 1px solid var(--border);
    border-radius: 8px; padding: 8px 12px; font-size: 13px;
  }
  .file-item .icon { font-size: 16px; }
  .file-item .fname { flex: 1; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }
  .file-item .size  { color: var(--muted); font-size: 11px; font-family: 'IBM Plex Mono', monospace; }
  .file-item .rm {
    cursor: pointer; color: var(--muted); font-size: 18px; line-height: 1;
    transition: color .15s; flex-shrink: 0;
  }
  .file-item .rm:hover { color: var(--danger); }

  /* Button */
  .btn {
    width: 100%; padding: 14px;
    background: linear-gradient(135deg, var(--accent) 0%, var(--accent2) 100%);
    color: white; font-weight: 700; font-size: 15px;
    border: none; border-radius: 12px; cursor: pointer;
    transition: all .2s; display: flex; align-items: center;
    justify-content: center; gap: 8px;
    box-shadow: 0 4px 20px rgba(59,130,246,.3);
    font-family: 'IBM Plex Sans', sans-serif;
    letter-spacing: .01em;
  }
  .btn:hover:not(:disabled) {
    transform: translateY(-1px);
    box-shadow: 0 6px 24px rgba(59,130,246,.45);
  }
  .btn:active:not(:disabled) { transform: translateY(0); }
  .btn:disabled { opacity: .45; cursor: not-allowed; box-shadow: none; transform: none; }

  /* Progress */
  .progress-wrap {
    height: 3px; background: var(--border); border-radius: 2px;
    overflow: hidden; margin-top: 12px; display: none;
  }
  .progress-bar {
    height: 100%; width: 0%;
    background: linear-gradient(90deg, var(--accent), var(--accent2));
    border-radius: 2px; transition: width .4s ease;
  }

  /* Status */
  .status-row {
    display: flex; align-items: center; gap: 10px;
    font-size: 13px; color: var(--muted); margin-top: 12px;
    min-height: 20px;
  }
  .dot {
    width: 7px; height: 7px; border-radius: 50%;
    background: var(--muted); flex-shrink: 0;
  }
  .dot.proc { background: var(--accent); animation: pulse 1s infinite; }
  .dot.ok   { background: var(--success); }
  .dot.err  { background: var(--danger); }
  @keyframes pulse { 0%,100%{opacity:1} 50%{opacity:.3} }

  /* Footer */
  footer {
    color: var(--muted); font-size: 12px; text-align: center;
    margin-top: 8px; line-height: 1.7;
  }

  /* Result box */
  .result-box {
    background: rgba(16,185,129,.08);
    border: 1px solid rgba(16,185,129,.3);
    border-radius: 10px; padding: 14px 16px;
    margin-top: 12px; display: none;
    align-items: center; gap: 12px; font-size: 13px;
  }
  .result-box.show { display: flex; }
  .result-icon { font-size: 24px; flex-shrink: 0; }
  .result-text { flex: 1; }
  .result-text strong { display: block; color: var(--success); font-size: 14px; margin-bottom: 2px; }

  @media (max-width: 400px) {
    body { padding: 16px 12px 32px; }
    .card-body { padding: 16px; }
  }
</style>
</head>
<body>
<div class="container">

  <div class="header">
    <div class="logo-wrap">
      <svg fill="none" stroke="currentColor" viewBox="0 0 24 24">
        <path stroke-linecap="round" stroke-linejoin="round" stroke-width="2"
          d="M9 17v-2m3 2v-4m3 4v-6m2 10H7a2 2 0 01-2-2V5a2 2 0 012-2h5.586a1 1 0 01.707.293l5.414 5.414a1 1 0 01.293.707V19a2 2 0 01-2 2z"/>
      </svg>
    </div>
    <h1>Bank Statement Converter</h1>
    <p class="subtitle">Konversi mutasi PDF ke Excel secara otomatis</p>
    <div class="banks">
      <span class="badge badge-bca">✦ BCA</span>
      <span class="badge badge-bri">✦ BRI</span>
      <span class="badge badge-mandiri">✦ Mandiri</span>
    </div>
  </div>

  <div class="card">
    <div class="card-header">Upload File PDF</div>
    <div class="card-body">
      <div class="drop-zone" id="dropZone">
        <input type="file" id="fileInput" multiple accept=".pdf">
        <div class="drop-icon">
          <svg fill="none" stroke="currentColor" viewBox="0 0 24 24">
            <path stroke-linecap="round" stroke-linejoin="round" stroke-width="2"
              d="M7 16a4 4 0 01-.88-7.903A5 5 0 1115.9 6L16 6a5 5 0 011 9.9M15 13l-3-3m0 0l-3 3m3-3v12"/>
          </svg>
        </div>
        <div class="drop-label">
          <strong>Klik atau seret file ke sini</strong><br>
          Bisa pilih 1–12 file PDF sekaligus (BCA / BRI / Mandiri)
        </div>
      </div>
      <div class="file-list" id="fileList"></div>
    </div>
  </div>

  <div class="card">
    <div class="card-header">Proses Konversi</div>
    <div class="card-body">
      <button class="btn" id="convertBtn" disabled onclick="startConvert()">
        <svg width="18" height="18" fill="none" stroke="currentColor" viewBox="0 0 24 24">
          <path stroke-linecap="round" stroke-linejoin="round" stroke-width="2.5"
            d="M4 16v1a3 3 0 003 3h10a3 3 0 003-3v-1m-4-8l-4-4m0 0L8 8m4-4v12"/>
        </svg>
        Generate Excel
      </button>
      <div class="progress-wrap" id="progressWrap">
        <div class="progress-bar" id="progressBar"></div>
      </div>
      <div class="status-row">
        <div class="dot" id="statusDot"></div>
        <span id="statusText">Pilih file PDF untuk memulai</span>
      </div>
      <div class="result-box" id="resultBox">
        <div class="result-icon">✅</div>
        <div class="result-text" id="resultText"></div>
      </div>
    </div>
  </div>

  <footer>
    Mendukung format e-statement BCA, BRI (IBIZ), dan Mandiri (Kopra)<br>
    Semua pemrosesan dilakukan di server — file tidak disimpan
  </footer>
</div>

<script>
const files = {};

const dz   = document.getElementById('dropZone');
const fi   = document.getElementById('fileInput');
const btn  = document.getElementById('convertBtn');
const list = document.getElementById('fileList');

// Drop zone
dz.addEventListener('dragover', e => { e.preventDefault(); dz.classList.add('dragover'); });
dz.addEventListener('dragleave', () => dz.classList.remove('dragover'));
dz.addEventListener('drop', e => {
  e.preventDefault(); dz.classList.remove('dragover');
  handleFiles([...e.dataTransfer.files]);
});
fi.addEventListener('change', e => handleFiles([...e.target.files]));
dz.addEventListener('click', e => { if (e.target === dz || e.target.closest('.drop-icon, .drop-label')) fi.click(); });

function fmtSize(bytes) {
  if (bytes < 1024) return bytes + ' B';
  if (bytes < 1048576) return (bytes/1024).toFixed(1) + ' KB';
  return (bytes/1048576).toFixed(1) + ' MB';
}

function handleFiles(newFiles) {
  newFiles.filter(f => f.name.toLowerCase().endsWith('.pdf')).forEach(f => {
    files[f.name] = f;
  });
  renderList();
}

function renderList() {
  list.innerHTML = '';
  const keys = Object.keys(files);
  keys.forEach(name => {
    const f = files[name];
    const div = document.createElement('div');
    div.className = 'file-item';
    div.innerHTML = `
      <span class="icon">📄</span>
      <span class="fname">${name}</span>
      <span class="size">${fmtSize(f.size)}</span>
      <span class="rm" data-name="${name}">×</span>`;
    list.appendChild(div);
  });
  list.querySelectorAll('.rm').forEach(el => {
    el.addEventListener('click', e => {
      delete files[e.target.dataset.name];
      renderList();
      updateBtn();
    });
  });
  updateBtn();
}

function updateBtn() {
  btn.disabled = Object.keys(files).length === 0;
}

function setStatus(state, text) {
  const dot = document.getElementById('statusDot');
  dot.className = 'dot ' + (state === 'proc' ? 'proc' : state === 'ok' ? 'ok' : state === 'err' ? 'err' : '');
  document.getElementById('statusText').textContent = text;
}

function setProgress(pct) {
  const w = document.getElementById('progressWrap');
  const b = document.getElementById('progressBar');
  w.style.display = pct > 0 ? 'block' : 'none';
  b.style.width = pct + '%';
}

async function startConvert() {
  const flist = Object.values(files);
  if (!flist.length) return;

  btn.disabled = true;
  document.getElementById('resultBox').classList.remove('show');
  setStatus('proc', `Mengirim ${flist.length} file ke server...`);
  setProgress(10);

  const fd = new FormData();
  flist.forEach(f => fd.append('files', f));

  try {
    setProgress(40);
    setStatus('proc', 'Mengekstrak transaksi dari PDF...');

    const res = await fetch('/convert', { method: 'POST', body: fd });
    setProgress(80);

    if (!res.ok) {
      const err = await res.text();
      throw new Error(err);
    }

    setProgress(95);
    setStatus('proc', 'Membuat file Excel...');

    const blob = await res.blob();
    const fname = res.headers.get('X-Filename') || 'Mutasi_Bank.xlsx';

    const url = URL.createObjectURL(blob);
    const a = document.createElement('a');
    a.href = url; a.download = fname; a.click();
    URL.revokeObjectURL(url);

    setProgress(100);
    setStatus('ok', `Selesai! File "${fname}" berhasil diunduh.`);
    const rb = document.getElementById('resultBox');
    document.getElementById('resultText').innerHTML =
      `<strong>${fname}</strong>File Excel siap — cek folder unduhan Anda.`;
    rb.classList.add('show');
    setTimeout(() => setProgress(0), 1500);

  } catch (err) {
    setStatus('err', 'Gagal: ' + err.message);
    setProgress(0);
  } finally {
    btn.disabled = false;
  }
}
</script>
</body>
</html>"""


@app.route('/')
def index():
    return render_template_string(HTML_TEMPLATE)


@app.route('/convert', methods=['POST'])
def convert():
    uploaded = request.files.getlist('files')
    if not uploaded:
        return 'Tidak ada file yang diunggah.', 400

    all_txs    = []
    period_set = set()
    bank_set   = set()
    saldo_awal = None
    saldo_akhir = 0
    mut_cr = mut_db = 0
    nama_akun = ''
    no_rek_set = set()

    for f in uploaded:
        raw = f.read()
        try:
            data = parse_pdf(raw)
        except Exception as e:
            return f'Error memproses {f.filename}: {str(e)}', 400

        if saldo_awal is None:
            saldo_awal = data['saldo_awal']
            nama_akun  = data['nama_akun']

        saldo_akhir = data['saldo_akhir']
        mut_cr     += data['mut_cr']
        mut_db     += data['mut_db']
        all_txs.extend(data['txs'])
        period_set.add(data['period'])
        bank_set.add(data['bank'])
        no_rek_set.add(data['no_rek'])

    merged = {
        'bank':       ' + '.join(sorted(bank_set)),
        'no_rek':     ' | '.join(sorted(no_rek_set)),
        'nama_akun':  nama_akun,
        'period':     ' | '.join(sorted(period_set)),
        'saldo_awal': saldo_awal or 0,
        'saldo_akhir':saldo_akhir,
        'mut_cr':     mut_cr,
        'mut_db':     mut_db,
        'txs':        all_txs,
    }

    excel = create_excel(merged)

    banks_str  = '_'.join(sorted(bank_set))
    fname      = f"Mutasi_{banks_str}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    mime       = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return send_file(
        excel,
        mimetype=mime,
        as_attachment=True,
        download_name=fname,
        headers={'X-Filename': fname}
    )


if __name__ == '__main__':
    app.run(debug=True, port=5000)
